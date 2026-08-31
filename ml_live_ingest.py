#!/usr/bin/env python3
"""Keep the live model's inputs CURRENT. TML is frozen at 2026-01-17; tennis-data is not.

THE CONSTRAINT, stated plainly. The point model is built from serve statistics, and the only free
source for those (TML) stopped updating on 2026-01-17 - 137 rows for the whole of 2026. Scoring an
August match on January ratings is not a stale-data annoyance, it is a different model.

tennis-data.co.uk publishes ATP results through 2026-08-29 with ranks and SET SCORES, but no serve
statistics. So the two halves of the model age differently:

    ELO          can be brought fully current from results alone            -> LIVE
    POINT MODEL  needs serve points, which stop in January                  -> STALE by 7 months

This ingests tennis-data 2026 (and WTA if present) into a results table so Elo can be replayed to
yesterday. It does NOT pretend to update the point model: a stale component that is labelled stale
can be weighted or dropped, while one that is silently stale corrupts everything downstream.
"""
import datetime as dt
import io
import sqlite3
import urllib.request
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
DB = HERE / "tennis_ace.sqlite"
UA = {"User-Agent": "Mozilla/5.0"}

con = sqlite3.connect(str(DB), timeout=60)
con.execute("PRAGMA busy_timeout=60000")
con.execute("""CREATE TABLE IF NOT EXISTS results_live(
    date TEXT, tour TEXT, tourney TEXT, surface TEXT, round TEXT, best_of INT,
    winner TEXT, loser TEXT, wrank REAL, lrank REAL, w_odds REAL, l_odds REAL,
    PRIMARY KEY (date, winner, loser))""")
con.commit()

# BOTH TOURS, ALL YEARS. TML is ATP-only, so building Elo from it left every WTA player unknown -
# which is exactly what produced 88 garbage bets at 0.500 in the first live scan. tennis-data
# carries both tours back to 2015, so Elo can cover WTA properly and stay current.
SRC = []
for _y in range(2015, 2027):
    SRC.append(("http://www.tennis-data.co.uk/%d/%d.xlsx" % (_y, _y), "ATP"))
    SRC.append(("http://www.tennis-data.co.uk/%dw/%d.xlsx" % (_y, _y), "WTA"))
tot = 0
for url, tour in SRC:
    raw = None
    for _try in (1, 2, 3):
        try:
            raw = urllib.request.urlopen(urllib.request.Request(url, headers=UA),
                                         timeout=240).read()
            break
        except Exception as e:                                          # noqa: BLE001
            err = str(e)[:40]
    if raw is None:
        print("   %-4s %s FETCH FAILED %s" % (tour, url[-14:], err))
        continue
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(it)]
    idx = {h: i for i, h in enumerate(hdr)}

    def g(row, *names):
        for nm in names:
            if nm in idx and idx[nm] < len(row):
                return row[idx[nm]]
        return None

    batch = []
    for row in it:
        w, l, d = g(row, "Winner"), g(row, "Loser"), g(row, "Date")
        if not w or not l or d is None:
            continue
        ds = d.date().isoformat() if isinstance(d, dt.datetime) else str(d)[:10]

        def f(x):
            try:
                return float(x)
            except (TypeError, ValueError):
                return None
        batch.append((ds, tour, str(g(row, "Tournament") or ""), str(g(row, "Surface") or ""),
                      str(g(row, "Round") or ""), int(g(row, "Best of") or 3), str(w), str(l),
                      f(g(row, "WRank")), f(g(row, "LRank")),
                      f(g(row, "PSW", "AvgW", "B365W")), f(g(row, "PSL", "AvgL", "B365L"))))
    if batch:
        con.executemany("INSERT OR IGNORE INTO results_live VALUES(?,?,?,?,?,?,?,?,?,?,?,?)", batch)
        con.commit()
        tot += len(batch)
    print("   %-4s %-10s %5d matches, latest %s"
          % (tour, url[-14:], len(batch), max(b[0] for b in batch) if batch else "-"))

n, lo, hi = con.execute("SELECT COUNT(*), MIN(date), MAX(date) FROM results_live").fetchone()
print("\nresults_live: %d matches, %s .. %s" % (n, lo, hi))
for t, c, mx in con.execute("SELECT tour, COUNT(*), MAX(date) FROM results_live GROUP BY tour"):
    print("   %-4s %5d  latest %s" % (t, c, mx))
srv_max = con.execute("SELECT MAX(date) FROM srv_pm").fetchone()[0]
print("\nSTALENESS: serve stats end %s, results end %s -> the point model is %d days behind"
      % (srv_max, hi,
         (dt.date.fromisoformat(hi) - dt.date.fromisoformat(srv_max)).days))
con.close()
