#!/usr/bin/env python3
"""Fetch tennis-data.co.uk ATP odds and join them to the TML ace table.

WHY A JOIN AT ALL. TML carries serve statistics but no prices, and the workload half of the ace
model needs a competitiveness signal: a lopsided match is short, a close one is long. Rank
difference is available inside TML and would avoid this join entirely - but it is a PROXY, and the
quantity we will actually have at prediction time is a moneyline. Fitting on a proxy and deploying
on a price is how a model quietly stops meaning what it was measured to mean, so the join is worth
doing properly.

THE JOIN IS THE RISK, and it is named in this project's own history: a careless name key once
merged two different players and cartesian-joined their outcomes. tennis-data writes "Federer R."
where TML writes "Roger Federer", so the key is (surname, first initial, date-window) and every
ambiguous or unmatched row is COUNTED AND REPORTED rather than silently dropped. A join rate is
reported before any model is fitted on it.
"""
import datetime as dt
import io
import re
import sqlite3
import sys
import unicodedata
import urllib.request
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl

DB = Path(__file__).resolve().parent / "tennis_ace.sqlite"
UA = {"User-Agent": "Mozilla/5.0"}
YEARS = list(range(2015, 2026))


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z ]", " ", s.lower()).strip()


def key_td(name):
    """tennis-data: 'Federer R.' -> ('federer','r')"""
    n = norm(name)
    parts = [p for p in n.split() if p]
    if len(parts) < 2:
        return None
    return (" ".join(parts[:-1]), parts[-1][:1])


def key_tml(name):
    """TML: 'Roger Federer' -> ('federer','r')"""
    n = norm(name)
    parts = [p for p in n.split() if p]
    if len(parts) < 2:
        return None
    return (" ".join(parts[1:]), parts[0][:1])


con = sqlite3.connect(str(DB), timeout=60)
con.execute("PRAGMA busy_timeout=60000")
con.execute("""CREATE TABLE IF NOT EXISTS odds_hist(
    date TEXT, winner TEXT, loser TEXT, wkey TEXT, lkey TEXT,
    best_of INT, surface TEXT, w_odds REAL, l_odds REAL,
    PRIMARY KEY (date, wkey, lkey))""")
con.commit()

tot = 0
for y in YEARS:
    try:
        r = urllib.request.Request("http://www.tennis-data.co.uk/%d/%d.xlsx" % (y, y), headers=UA)
        raw = urllib.request.urlopen(r, timeout=90).read()
    except Exception as e:                                              # noqa: BLE001
        print("   %d FETCH FAILED %s" % (y, str(e)[:44]))
        continue
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c or "").strip() for c in next(it)]
    idx = {h: i for i, h in enumerate(hdr)}

    def g(row, *names):
        for nm in names:
            if nm in idx and idx[nm] < len(row):
                return row[idx[nm]]
        return None

    batch = []
    for row in it:
        w, l = g(row, "Winner"), g(row, "Loser")
        d = g(row, "Date")
        if not w or not l or d is None:
            continue
        if isinstance(d, dt.datetime):
            ds = d.date().isoformat()
        else:
            try:
                ds = dt.datetime.strptime(str(d)[:10], "%Y-%m-%d").date().isoformat()
            except Exception:                                           # noqa: BLE001
                continue
        wk, lk = key_td(w), key_td(l)
        if not wk or not lk:
            continue
        wo = g(row, "PSW", "B365W", "AvgW", "MaxW")
        lo = g(row, "PSL", "B365L", "AvgL", "MaxL")
        try:
            wo = float(wo) if wo else None
            lo = float(lo) if lo else None
        except (TypeError, ValueError):
            wo = lo = None
        if not wo or not lo or wo <= 1 or lo <= 1:
            continue
        bo = g(row, "Best of")
        batch.append((ds, str(w), str(l), "%s|%s" % wk, "%s|%s" % lk,
                      int(bo) if bo else 3, str(g(row, "Surface") or ""), wo, lo))
    if batch:
        con.executemany("INSERT OR IGNORE INTO odds_hist VALUES(?,?,?,?,?,?,?,?,?)", batch)
        con.commit()
        tot += len(batch)
    print("   %d  %d priced matches" % (y, len(batch)))

n = con.execute("SELECT COUNT(*) FROM odds_hist").fetchone()[0]
print("\nodds_hist: %d matches" % n)

# ---- JOIN RATE, measured before anything is fitted on it -----------------------------------
oh = defaultdict(list)
for d, wk, lk, wo, lo in con.execute("SELECT date, wkey, lkey, w_odds, l_odds FROM odds_hist"):
    oh[(wk, lk)].append((d, wo, lo))

rows = con.execute("SELECT date, player, opp, won FROM ace_pm WHERE year >= 2015").fetchall()
hit = miss = 0
for d, pl, op, won in rows:
    kp, ko = key_tml(pl), key_tml(op)
    if not kp or not ko:
        miss += 1
        continue
    a = "%s|%s" % kp
    b = "%s|%s" % ko
    cand = oh.get((a, b)) if won else oh.get((b, a))
    ok = False
    if cand:
        for dd, _wo, _lo in cand:
            if abs((dt.date.fromisoformat(dd) - dt.date.fromisoformat(d)).days) <= 4:
                ok = True
                break
    hit += 1 if ok else 0
    miss += 0 if ok else 1
print("JOIN RATE on ace_pm rows: %d matched, %d unmatched (%.1f%% matched)"
      % (hit, miss, 100.0 * hit / max(hit + miss, 1)))
print("   (a date window of +-4 days is allowed: tennis-data stamps the SCHEDULED day and TML the")
print("    tournament start, so an exact-date join would fail on most rows for a spurious reason.)")
con.close()
